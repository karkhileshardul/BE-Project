import cognitive_face as CF
from global_variables import personGroupId
import sys
import os, urllib
import sqlite3
from openpyxl.utils import get_column_letter
from openpyxl import workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import column_index_from_string
#from openpyxl.utils import get_column_letter, cell, column_index_from_string
import time
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports.xlsx")
sheet = wb.get_sheet_by_name('Cse15')

def getDateColumn():
	for i in range(1, len(sheet.rows[0]) + 1):
		col = get_column_letter(i)
		if sheet.cell('%s%s'% (col,'1')).value == currentDate:
			return col
			
			
#BASE_URL = 'https://southeastasia.api.cognitive.microsoft.com/face/v1.0'
#CF.BaseUrl.set(BASE_URL)
#Key = 'a7073ac5c5af4f4d9f6c1a609efc3a56'
#CF.Key.set(Key)

BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0'
CF.BaseUrl.set(BASE_URL)
Key = '2c9ae23bb71242cca1eb17e10322fc6a'
CF.Key.set(Key)


connect = connect = sqlite3.connect("Face-DataBase")
c = connect.cursor()

attend = [0 for i in range(60)]	


def some_function(imageFolder):
	for filename in os.listdir(imageFolder):
		if filename.endswith(".jpg"):
	#        	print(filename)
	#        	filename = urllib.pathname2url(os.path.join(imageFolder, filename))
	#		detecting_face()
			os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
			i=1
			while(i<17):
				if(i==1):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="1.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==2):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="2.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==3):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="3.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==4):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="4.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==5):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="5.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==6):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="6.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==7):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="7.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==8):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="8.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==9):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="9.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==10):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="10.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==11):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="11.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==12):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="12.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==13):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="13.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==14):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="14.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==15):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="15.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)
				if(i==16):
					os.chdir("/home/shardul/BE PROJECT/Cropped_faces/")
					filename="16.jpg"
					print(filename)
					res = CF.face.detect(filename)
					if len(res) != 1:
						print "No face detected in image"
						continue
					faceIds = []
					for face in res:
						faceIds.append(face['faceId'])
					res = CF.face.identify(faceIds, personGroupId)
					print filename
					print res
					for face  in res:
						if not face['candidates']:
							print "Unknown"
						else:
							personId = face['candidates'][0]['personId']
							c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
							row = c.fetchone()
							attend[int(row[0])] += 1
							print row[1] + " recognized"
					time.sleep(3)

				i=i+1
		

for row in range(2, len(sheet.columns[0]) + 1):
	rn = sheet.cell('A%s'% row).value
	if rn is not None:
		rn = rn[-2:]
		if attend[int(rn)] != 0:
			col = getDateColumn()
			sheet['%s%s' % (col, str(row))] = 1

wb.save(filename = "reports.xlsx")	 	

#currentDir = os.path.dirname(os.path.abspath(__file__))
#directory = os.path.join(currentDir, 'Cropped_faces')
#for filename in os.listdir(directory):
#	if filename.endswith(".jpg"):
#		filename = urllib.pathname2url(os.path.join(directory, filename))
#		res = CF.face.detect(filename)
#		if len(res) != 1:
#			print "No face detected."
#			continue
#		some_function(directory)				
'''
		faceIds = []
		for face in res:
			faceIds.append(face['faceId'])
		res = CF.face.identify(faceIds, personGroupId)
		print filename
		print res
		for face  in res:
			if not face['candidates']:
				print "Unknown"
			else:
				personId = face['candidates'][0]['personId']
				c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
				row = c.fetchone()
				attend[int(row[0])] += 1
				print row[1] + " recognized"
		time.sleep(6)
	
	
for row in range(2, len(sheet.columns[0]) + 1):
	rn = sheet.cell('A%s'% row).value
	if rn is not None:
		rn = rn[-2:]
		if attend[int(rn)] != 0:
			col = getDateColumn()
			sheet['%s%s' % (col, str(row))] = 1

wb.save(filename = "reports.xlsx")	 	
'''

#currentDir = os.path.dirname(os.path.abspath(__file__))
#filename = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(filename)
#faceIds = []
#for face in res:
 #   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
#print res
